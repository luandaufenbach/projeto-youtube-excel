import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep 
# 1 entrar na planilha e extrair cpf do cliente

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/') #fora do for para nao ficar abrindo varias vezes

for linha in pagina_clientes.iter_rows(min_row=2,values_only=True): #comeca da linha 2/ retorna somente valores
    nome,valor,cpf,vencimento = linha
    #entrar no site https://consultcpf-devaprender.netlify.app/ e usar o cpf da planilha para pesquisar 
    #o status do pagamento
    sleep(5) #delay de 5s
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    #3 - verificar se esta "em dia" ou "atrasado"
    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)
    
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    if status.text == 'em dia':
        #4 - se estiver "em dia", pegar a data do pagamento e o metodo de pagamento
        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
        data_pagamento_limpo = data_pagamento.text.split()[3] #mostra somente a data
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3] #mostra somente o metodo
    
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome,valor,cpf,vencimento,'em dia',
        data_pagamento_limpo,metodo_pagamento_limpo]) #quando em dia mostra data do pagamento e o metodo do pagamento
        
        planilha_fechamento.save('planilha fechamento.xlsx')

    else:
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        #5 - caso contrario (atrasado), colocar o status como pendente
        pagina_fechamento.append([nome, valor, cpf, vencimento,'pendente']) #quando nao estiver em dia mostra somente o pendente 
        planilha_fechamento.save('planilha fechamento.xlsx')        
    
